"""
特殊数据类型解析器
支持表格、数学公式、图表等特殊元素的专业解析
"""
import asyncio
import time
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import xml.etree.ElementTree as ET

import pandas as pd
import numpy as np
from PIL import Image
import cv2

from src.data.parsers.base import BaseParser, ParseResult, ParsedElement
from src.data.document_validator import DocumentInfo, DocumentType
from src.utils.logger import get_logger
from config.settings import settings

logger = get_logger("data.parsers.special")


class TableExtractionParser(BaseParser):
    """表格提取专业解析器"""
    
    def __init__(self):
        super().__init__("TableExtractionParser")
        self.supported_types = [
            DocumentType.PDF, DocumentType.DOCX, DocumentType.PPTX,
            DocumentType.HTML, DocumentType.XLSX, DocumentType.XLS
        ]
    
    async def parse(self, file_path: Path, doc_info: DocumentInfo) -> ParseResult:
        """专业表格提取"""
        start_time = time.time()
        
        try:
            elements = []
            
            if doc_info.file_type in [DocumentType.XLSX, DocumentType.XLS]:
                elements = await self._extract_excel_tables(file_path)
            elif doc_info.file_type == DocumentType.PDF:
                elements = await self._extract_pdf_tables(file_path)
            elif doc_info.file_type == DocumentType.HTML:
                elements = await self._extract_html_tables(file_path)
            elif doc_info.file_type in [DocumentType.DOCX]:
                elements = await self._extract_docx_tables(file_path)
            
            processing_time = time.time() - start_time
            
            return ParseResult(
                success=True,
                elements=elements,
                metadata={
                    'parser': self.name,
                    'extraction_method': 'specialized_table_extraction',
                    'tables_found': len(elements)
                },
                processing_time=processing_time
            )
            
        except Exception as e:
            return self._handle_error(e, file_path)
    
    async def _extract_excel_tables(self, file_path: Path) -> List[ParsedElement]:
        """提取Excel表格"""
        elements = []
        
        try:
            excel_file = pd.ExcelFile(str(file_path))
            
            for sheet_index, sheet_name in enumerate(excel_file.sheet_names):
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                if df.empty:
                    continue
                
                # 检测表格结构
                table_info = self._analyze_table_structure(df)
                
                # 转换为增强的Markdown格式
                enhanced_markdown = self._create_enhanced_table_markdown(df, table_info)
                
                table_element = self._create_table_element(
                    content=enhanced_markdown,
                    metadata={
                        'sheet_index': sheet_index,
                        'sheet_name': sheet_name,
                        'rows': len(df),
                        'cols': len(df.columns),
                        'table_type': table_info['type'],
                        'has_header': table_info['has_header'],
                        'numeric_columns': table_info['numeric_columns'],
                        'extraction_method': 'pandas_enhanced'
                    },
                    position={'sheet': sheet_index, 'sheet_name': sheet_name}
                )
                elements.append(table_element)
                
        except Exception as e:
            logger.warning(f"Excel表格提取失败: {e}")
        
        return elements
    
    async def _extract_pdf_tables(self, file_path: Path) -> List[ParsedElement]:
        """提取PDF表格"""
        elements = []
        
        try:
            import fitz  # PyMuPDF
            
            doc = fitz.open(str(file_path))
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                
                # 使用PyMuPDF的表格检测
                tables = page.find_tables()
                
                for table_index, table in enumerate(tables):
                    table_data = table.extract()
                    
                    if table_data:
                        df = pd.DataFrame(table_data[1:], columns=table_data[0])
                        table_info = self._analyze_table_structure(df)
                        
                        enhanced_markdown = self._create_enhanced_table_markdown(df, table_info)
                        
                        table_element = self._create_table_element(
                            content=enhanced_markdown,
                            metadata={
                                'page_number': page_num + 1,
                                'table_index': table_index,
                                'rows': len(df),
                                'cols': len(df.columns),
                                'table_type': table_info['type'],
                                'extraction_method': 'pymupdf_enhanced'
                            },
                            position={'page': page_num + 1, 'table': table_index}
                        )
                        elements.append(table_element)
            
            doc.close()
            
        except Exception as e:
            logger.warning(f"PDF表格提取失败: {e}")
        
        return elements
    
    async def _extract_html_tables(self, file_path: Path) -> List[ParsedElement]:
        """提取HTML表格"""
        elements = []
        
        try:
            tables = pd.read_html(str(file_path))
            
            for table_index, df in enumerate(tables):
                table_info = self._analyze_table_structure(df)
                enhanced_markdown = self._create_enhanced_table_markdown(df, table_info)
                
                table_element = self._create_table_element(
                    content=enhanced_markdown,
                    metadata={
                        'table_index': table_index,
                        'rows': len(df),
                        'cols': len(df.columns),
                        'table_type': table_info['type'],
                        'extraction_method': 'pandas_html'
                    },
                    position={'table': table_index}
                )
                elements.append(table_element)
                
        except Exception as e:
            logger.warning(f"HTML表格提取失败: {e}")
        
        return elements
    
    async def _extract_docx_tables(self, file_path: Path) -> List[ParsedElement]:
        """提取DOCX表格"""
        elements = []
        
        try:
            from docx import Document
            
            doc = Document(str(file_path))
            
            for table_index, table in enumerate(doc.tables):
                table_data = []
                for row in table.rows:
                    row_data = [cell.text.strip() for cell in row.cells]
                    table_data.append(row_data)
                
                if table_data:
                    df = pd.DataFrame(table_data[1:], columns=table_data[0])
                    table_info = self._analyze_table_structure(df)
                    
                    enhanced_markdown = self._create_enhanced_table_markdown(df, table_info)
                    
                    table_element = self._create_table_element(
                        content=enhanced_markdown,
                        metadata={
                            'table_index': table_index,
                            'rows': len(df),
                            'cols': len(df.columns),
                            'table_type': table_info['type'],
                            'extraction_method': 'python_docx_enhanced'
                        },
                        position={'table': table_index}
                    )
                    elements.append(table_element)
                    
        except Exception as e:
            logger.warning(f"DOCX表格提取失败: {e}")
        
        return elements
    
    def _analyze_table_structure(self, df: pd.DataFrame) -> Dict[str, Any]:
        """分析表格结构"""
        info = {
            'type': 'data_table',
            'has_header': True,
            'numeric_columns': [],
            'categorical_columns': [],
            'date_columns': []
        }
        
        # 检测数值列
        for col in df.columns:
            try:
                pd.to_numeric(df[col], errors='raise')
                info['numeric_columns'].append(col)
            except:
                # 检测日期列
                try:
                    pd.to_datetime(df[col], errors='raise')
                    info['date_columns'].append(col)
                except:
                    info['categorical_columns'].append(col)
        
        # 判断表格类型
        if len(info['numeric_columns']) > len(info['categorical_columns']):
            info['type'] = 'numeric_table'
        elif len(info['date_columns']) > 0:
            info['type'] = 'time_series_table'
        
        return info
    
    def _create_enhanced_table_markdown(self, df: pd.DataFrame, table_info: Dict[str, Any]) -> str:
        """创建增强的表格Markdown"""
        # 基础Markdown表格
        markdown = df.to_markdown(index=False)
        
        # 添加表格元信息
        metadata_lines = [
            f"<!-- 表格类型: {table_info['type']} -->",
            f"<!-- 行数: {len(df)}, 列数: {len(df.columns)} -->",
        ]
        
        if table_info['numeric_columns']:
            metadata_lines.append(f"<!-- 数值列: {', '.join(table_info['numeric_columns'])} -->")
        
        if table_info['date_columns']:
            metadata_lines.append(f"<!-- 日期列: {', '.join(table_info['date_columns'])} -->")
        
        enhanced_markdown = "\n".join(metadata_lines) + "\n\n" + markdown
        
        return enhanced_markdown


class FormulaExtractionParser(BaseParser):
    """数学公式提取解析器"""
    
    def __init__(self):
        super().__init__("FormulaExtractionParser")
        self.supported_types = [
            DocumentType.PDF, DocumentType.DOCX, DocumentType.XLSX,
            DocumentType.HTML, DocumentType.MD
        ]
        
        # 数学符号模式
        self.math_patterns = [
            r'[∑∏∫∮∯∰∱∲∳]',  # 求和、积分符号
            r'[√∛∜]',  # 根号
            r'[∞∝∈∉∋∌∅∆∇]',  # 数学符号
            r'[≤≥≠≈≡≢≅≆≇≈≉≊≋≌≍≎≏]',  # 比较符号
            r'[αβγδεζηθικλμνξοπρστυφχψω]',  # 希腊字母
            r'[ΑΒΓΔΕΖΗΘΙΚΛΜΝΞΟΠΡΣΤΥΦΧΨΩ]',  # 大写希腊字母
            r'\$[^$]+\$',  # LaTeX行内公式
            r'\$\$[^$]+\$\$',  # LaTeX块公式
            r'\\[a-zA-Z]+\{[^}]*\}',  # LaTeX命令
        ]
        
        self.compiled_patterns = [re.compile(pattern) for pattern in self.math_patterns]
    
    async def parse(self, file_path: Path, doc_info: DocumentInfo) -> ParseResult:
        """提取数学公式"""
        start_time = time.time()
        
        try:
            elements = []
            
            if doc_info.file_type == DocumentType.PDF:
                elements = await self._extract_pdf_formulas(file_path)
            elif doc_info.file_type == DocumentType.DOCX:
                elements = await self._extract_docx_formulas(file_path)
            elif doc_info.file_type == DocumentType.XLSX:
                elements = await self._extract_excel_formulas(file_path)
            elif doc_info.file_type == DocumentType.HTML:
                elements = await self._extract_html_formulas(file_path)
            elif doc_info.file_type == DocumentType.MD:
                elements = await self._extract_markdown_formulas(file_path)
            
            processing_time = time.time() - start_time
            
            return ParseResult(
                success=True,
                elements=elements,
                metadata={
                    'parser': self.name,
                    'extraction_method': 'formula_extraction',
                    'formulas_found': len(elements)
                },
                processing_time=processing_time
            )
            
        except Exception as e:
            return self._handle_error(e, file_path)
    
    async def _extract_pdf_formulas(self, file_path: Path) -> List[ParsedElement]:
        """从PDF提取公式"""
        elements = []
        
        try:
            import fitz
            
            doc = fitz.open(str(file_path))
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                text = page.get_text()
                
                formulas = self._find_formulas_in_text(text)
                
                for formula_index, formula in enumerate(formulas):
                    formula_element = self._create_formula_element(
                        content=formula['content'],
                        metadata={
                            'page_number': page_num + 1,
                            'formula_index': formula_index,
                            'formula_type': formula['type'],
                            'confidence': formula['confidence'],
                            'extraction_method': 'regex_pattern'
                        },
                        position={'page': page_num + 1, 'formula': formula_index}
                    )
                    elements.append(formula_element)
            
            doc.close()
            
        except Exception as e:
            logger.warning(f"PDF公式提取失败: {e}")
        
        return elements
    
    async def _extract_excel_formulas(self, file_path: Path) -> List[ParsedElement]:
        """从Excel提取公式"""
        elements = []
        
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(str(file_path))
            
            for sheet_index, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                
                formula_index = 0
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # 公式类型
                            formula_element = self._create_formula_element(
                                content=f"{cell.coordinate}: {cell.value}",
                                metadata={
                                    'sheet_index': sheet_index,
                                    'sheet_name': sheet_name,
                                    'cell_coordinate': cell.coordinate,
                                    'formula_index': formula_index,
                                    'extraction_method': 'openpyxl'
                                },
                                position={'sheet': sheet_index, 'cell': cell.coordinate}
                            )
                            elements.append(formula_element)
                            formula_index += 1
            
        except Exception as e:
            logger.warning(f"Excel公式提取失败: {e}")
        
        return elements
    
    async def _extract_markdown_formulas(self, file_path: Path) -> List[ParsedElement]:
        """从Markdown提取公式"""
        elements = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            formulas = self._find_formulas_in_text(content)
            
            for formula_index, formula in enumerate(formulas):
                formula_element = self._create_formula_element(
                    content=formula['content'],
                    metadata={
                        'formula_index': formula_index,
                        'formula_type': formula['type'],
                        'confidence': formula['confidence'],
                        'extraction_method': 'regex_pattern'
                    },
                    position={'formula': formula_index}
                )
                elements.append(formula_element)
                
        except Exception as e:
            logger.warning(f"Markdown公式提取失败: {e}")
        
        return elements
    
    def _find_formulas_in_text(self, text: str) -> List[Dict[str, Any]]:
        """在文本中查找公式"""
        formulas = []
        
        for pattern_index, pattern in enumerate(self.compiled_patterns):
            matches = pattern.finditer(text)
            
            for match in matches:
                formula = {
                    'content': match.group(),
                    'type': self._classify_formula_type(match.group()),
                    'confidence': 0.8,  # 基于模式匹配的置信度
                    'start': match.start(),
                    'end': match.end()
                }
                formulas.append(formula)
        
        # 去重和排序
        formulas = self._deduplicate_formulas(formulas)
        formulas.sort(key=lambda x: x['start'])
        
        return formulas
    
    def _classify_formula_type(self, formula: str) -> str:
        """分类公式类型"""
        if any(symbol in formula for symbol in ['∑', '∏']):
            return 'summation'
        elif any(symbol in formula for symbol in ['∫', '∮', '∯', '∰']):
            return 'integral'
        elif any(symbol in formula for symbol in ['√', '∛', '∜']):
            return 'root'
        elif formula.startswith('$$') and formula.endswith('$$'):
            return 'latex_block'
        elif formula.startswith('$') and formula.endswith('$'):
            return 'latex_inline'
        elif any(symbol in formula for symbol in ['α', 'β', 'γ', 'δ']):
            return 'greek_letters'
        else:
            return 'general_math'
    
    def _deduplicate_formulas(self, formulas: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """去除重复的公式"""
        seen = set()
        unique_formulas = []
        
        for formula in formulas:
            content = formula['content']
            if content not in seen:
                seen.add(content)
                unique_formulas.append(formula)
        
        return unique_formulas
    
    async def _extract_docx_formulas(self, file_path: Path) -> List[ParsedElement]:
        """从DOCX提取公式（占位符）"""
        # 这里可以扩展DOCX公式提取逻辑
        return []
    
    async def _extract_html_formulas(self, file_path: Path) -> List[ParsedElement]:
        """从HTML提取公式（占位符）"""
        # 这里可以扩展HTML公式提取逻辑
        return []
